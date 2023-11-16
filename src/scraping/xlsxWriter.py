from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from scraping.utils import (
    XLSX_HEADERS,
    getReportsPath,
    getExcelPath,
    isFileCreated,
    createReportsDir,
    getCurrentTime,
)


def writeXlsxFile(formatedData: list[dict], username: str):
    REPORTS_PATH = getReportsPath(username)
    EXCEL_PATH = getExcelPath(REPORTS_PATH)

    currentTime = getCurrentTime()

    # Checks if the file is created
    if not isFileCreated(REPORTS_PATH):
        createReportsDir(REPORTS_PATH)

    if isFileCreated(EXCEL_PATH):
        # Load the existing workbook
        workbook = load_workbook(filename=EXCEL_PATH)

        # Create a new sheet with a unique name
        sheet_name = f"{currentTime}_Cesusc"
        worksheet = workbook.create_sheet(sheet_name)

    else:
        # Create a new workbook and worksheet with openpyxl
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = f"{currentTime}_Cesusc"

    # Create all the headers
    for col, header in enumerate(XLSX_HEADERS, start=1):
        worksheet.cell(row=1, column=col, value=header)

    # Resize the columns
    worksheet = resizeXslxColumns(worksheet, formatedData)

    # starts from the second row because the first row is the headers
    row = 2
    for subject in formatedData:
        # the first value is the subject name || tranforming the dict keys into a string
        worksheet.cell(row=row, column=1, value=str(list(subject.keys())))

        # get all the items from the dict and interate over them
        for key, listOfSubjectData in subject.items():
            for col, value in enumerate(listOfSubjectData, start=2):
                worksheet.cell(row=row, column=col, value=value)
        row += 1

    # Save the workbook
    workbook.save(filename=EXCEL_PATH)


def resizeXslxColumns(worksheet: Workbook, formatedData: list[dict]) -> Workbook:
    """This function resizes the columns of the xlsx file base on the size of the headers"""

    # Adjust the width of the columns based on the headers
    for i, column_cells in enumerate(
        worksheet.iter_cols(min_row=1, max_row=1), start=1
    ):
        header_cell = column_cells[0]

        # Checks if the header is the subject name
        if header_cell.value == "Nome da Materia":
            # Initialize longestKeyLength to 0
            longestKeyLength = 0
            # iterate the list of dicts with the subject data

            for data in formatedData:
                # Update longestKeyLength if necessary
                longestKeyLength = max(
                    longestKeyLength, max(len(key) for key in data.keys())
                )

            # Adjust the width of the column
            worksheet.column_dimensions[get_column_letter(i)].width = (
                longestKeyLength + 5
            )
            continue

        max_length = len(str(header_cell.value))
        adjusted_width = max_length + 2

        worksheet.column_dimensions[get_column_letter(i)].width = adjusted_width

    return worksheet
